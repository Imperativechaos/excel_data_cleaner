import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def clean_and_format_excel(input_file, output_file=None):
    """
    Cleans Excel data: removes empty rows/columns, standardizes headers,
    fills missing values, and applies professional formatting.
    """
    if output_file is None:
        output_file = f"cleaned_{os.path.basename(input_file)}"
    
    try:
        # Read Excel with pandas
        df = pd.read_excel(input_file)
        
        # Basic cleaning
        df = df.dropna(how='all')               # Drop fully empty rows
        df.columns = df.columns.str.strip().str.title()  # Clean headers
        df.fillna("N/A", inplace=True)          # Fill missing with placeholder
        
        # Save cleaned data back to Excel
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Apply formatting with openpyxl
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Header row bold + colored
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_file)
        print(f"Cleaned and formatted file saved as: {output_file}")
        return output_file
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return None

# Example usage (comment out or remove for client delivery)
# clean_and_format_excel("messy_sales_data.xlsx")
